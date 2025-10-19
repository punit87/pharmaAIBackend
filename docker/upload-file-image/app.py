import json
import re
import chardet
import base64
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from requests_toolbelt.multipart import decoder


def add_cors_headers(response):
    print(f"Adding CORS headers to the response...")
    response["headers"] = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type, Authorization"
    }
    return response


def lambda_handler(event, context):
    """AWS Lambda handler function."""
    try:
        # Debug: Log the incoming event
        print("Incoming event:", event)

        # Parse the content type and body
        content_type = event["headers"].get("content-type") or event["headers"].get("Content-Type")
        if not content_type or not content_type.startswith("multipart/form-data"):
            response = {
                "statusCode": 400,
                "body": json.dumps({"error": "Invalid content type"})
            }
            return add_cors_headers(response)

        # Decode multipart form data
        body = event["body"]
        is_base64_encoded = event.get("isBase64Encoded", False)
        if is_base64_encoded:
            body = base64.b64decode(body)
        else:
            body = body.encode("utf-8")

        multipart_data = decoder.MultipartDecoder(body, content_type)
        files = []
        for part in multipart_data.parts:
            content_disposition = part.headers.get(b"Content-Disposition", b"").decode()
            match = re.search(r'filename="(.+?)"', content_disposition)
            if match:
                filename = match.group(1)
                files.append({"filename": filename, "content": part.content})

        if not files:
            response = {
                "statusCode": 400,
                "body": json.dumps({"error": "No files uploaded"})
            }
            return add_cors_headers(response)

        # Process the first file (for simplicity)
        file = files[0]
        filename = file["filename"]
        raw_bytes = file["content"]

        print(f"Filename: {filename}, Size: {len(raw_bytes)} bytes")

        # Detect file encoding
        encoding_result = chardet.detect(raw_bytes)
        encoding = encoding_result['encoding']
        confidence = encoding_result['confidence']
        print(f"Detected encoding: {encoding} (confidence: {confidence})")

        if not encoding or confidence < 0.5:
            try:
                file_content = raw_bytes.decode("utf-8")
                print("Decoded using utf-8")
            except UnicodeDecodeError:
                raise Exception("Failed to decode file content as utf-8")
        else:
            file_content = raw_bytes.decode(encoding)

        # Process based on file type
        if filename.endswith(('.txt', '.text')):
            file_type = 'text'
            sections = re.split(r'\n\s*\n', file_content.strip())
            parsed_content = []
            for section in sections:
                lines = section.strip().split("\n")
                if lines:
                    subheading = lines[0].strip()
                    table_content = "\n".join(lines[1:]).strip()
                    parsed_content.append({"subheading": subheading, "table_content": table_content})
        elif filename.endswith(('.csv', '.tsv', '.xlsx', '.xls')):
            file_type = 'excel'

            # Handle Excel files
            if filename.endswith(('.xlsx', '.xls')):
                try:
                    workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True)
                    sheet = workbook.active
                    data = [[cell.value for cell in row] for row in sheet.iter_rows()]
                    df = pd.DataFrame(data)
                except Exception as e:
                    raise Exception(f"Failed to process Excel file: {e}")
            else:
                # Handle CSV/TSV files
                try:
                    delimiter = ',' if filename.endswith('.csv') else '\t'
                    df = pd.read_csv(BytesIO(raw_bytes), sep=delimiter, header=None, dtype=str)
                except Exception as e:
                    raise Exception(f"Failed to process CSV/TSV file: {e}")

            # Parse DataFrame into sections
            parsed_content = []
            current_subheading = None
            current_content = ""
            for index, row in df.iterrows():
                subheading = row.iloc[0] if len(row) > 0 else None
                content = row.iloc[1] if len(row) > 1 else None

                if pd.notna(subheading):
                    if current_subheading:
                        parsed_content.append({"subheading": current_subheading, "table_content": current_content.strip()})
                    current_subheading = subheading
                    current_content = str(content) + "\n" if pd.notna(content) else "\n"
                elif current_subheading and pd.notna(content):
                    current_content += str(content) + "\n"
            if current_subheading:
                parsed_content.append({"subheading": current_subheading, "table_content": current_content.strip()})
        else:
            response = {
                "statusCode": 400,
                "body": json.dumps({"error": "Unsupported file type"})
            }
            return add_cors_headers(response)

        response = {
            "statusCode": 200,
            "body": json.dumps({"results": [{"content": parsed_content}]})
        }
        return add_cors_headers(response)

    except Exception as e:
        print(f"Error processing file: {e}")
        response = {
            "statusCode": 500,
            "body": json.dumps({"error": f"Error processing file: {e}"})
        }
        return add_cors_headers(response)
